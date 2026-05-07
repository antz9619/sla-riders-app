import streamlit as st
import pandas as pd
import numpy as np
import re
from datetime import datetime, date, timedelta
import plotly.express as px
import plotly.graph_objects as go
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import BarChart, LineChart, Reference, PieChart
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint


# ============================================================
# 1. FERIADOS Y CÁLCULO DE DÍAS HÁBILES (SIN CAMBIOS)
# ============================================================
FERIADOS_INAMOVIBLES = {
    (1,1), (3,24), (4,2), (5,1), (5,25), (6,20), (7,9), (12,8), (12,25)
}

FERIADOS_VARIABLES = {
    2025: {(2,3), (2,4), (3,3), (3,4), (4,18), (4,19), (6,16), (8,15), (10,12), (11,17)},
    2026: {(2,16), (2,17), (4,3), (4,6), (6,15), (8,17), (10,12), (11,23), (12,24), (12,31)},
}

def es_feriado(fecha):
    if isinstance(fecha, datetime):
        fecha = fecha.date()
    par = (fecha.month, fecha.day)
    if par in FERIADOS_INAMOVIBLES:
        return True
    return par in FERIADOS_VARIABLES.get(fecha.year, set())

def es_dia_habil(fecha):
    if isinstance(fecha, datetime):
        fecha = fecha.date()
    return fecha.weekday() < 5 and not es_feriado(fecha)

def construir_feriados_np(anios):
    feriados = []
    for anio in anios:
        for mes, dia in FERIADOS_INAMOVIBLES:
            try:
                feriados.append(date(anio, mes, dia))
            except ValueError:
                pass
        for mes, dia in FERIADOS_VARIABLES.get(anio, set()):
            try:
                feriados.append(date(anio, mes, dia))
            except ValueError:
                pass
    return np.array(sorted(set(feriados)), dtype='datetime64[D]')

# ============================================================
# 2. NORMALIZAR GUÍA (SIN CAMBIOS)
# ============================================================
def pad_guia(guia):
    if pd.isna(guia):
        return guia
    match = re.match(r'^(.*-P-)(\d+)$', str(guia))
    if match:
        prefijo = match.group(1)
        numeros = match.group(2)
        return f"{prefijo}{numeros.zfill(8)}"
    return guia

# ============================================================
# 3. HOJA "Evolución Mensual SLA" (SIN CAMBIOS)
# ============================================================
def escribir_hoja_evolucion(writer, sla_global, evo_zonas, sla_semanal=None):
    C_HEADER_BLUE = "00468c"
    C_BLUE = "0072ce"
    C_TEAL = "00b0b0"
    C_ORANGE = "ff9100"
    C_WHITE = "ffffff"
    C_LIGHT = "f2f5fa"

    wb = writer.book
    ws = wb.create_sheet("Evolución Mensual SLA")

    align_center = Alignment(horizontal='center', vertical='center')
    font_normal = Font(name='Segoe UI', size=10)
    font_bold = Font(name='Segoe UI', size=10, bold=True)
    font_section = Font(name='Segoe UI', size=11, bold=True, color=C_WHITE)
    font_header = Font(name='Segoe UI', size=9, bold=True, color=C_WHITE)
    fill_section = PatternFill(start_color=C_HEADER_BLUE, end_color=C_HEADER_BLUE, fill_type='solid')
    fill_header = PatternFill(start_color=C_BLUE, end_color=C_BLUE, fill_type='solid')
    fill_light = PatternFill(start_color=C_LIGHT, end_color=C_LIGHT, fill_type='solid')

    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 12

    def escribir_tabla(ws, start_row, titulo, df_tabla, color_header,
                       incluir_total=True, col_etiqueta='label', texto_encabezado='Mes'):
        row = start_row
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        cell = ws.cell(row=row, column=1, value=titulo)
        cell.font = font_section
        cell.fill = PatternFill(start_color=color_header, end_color=color_header, fill_type='solid')
        row += 1

        headers = [texto_encabezado, 'Gestionables', 'A Tiempo', '% SLA']
        for i, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=i, value=h)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center
        row += 1

        if df_tabla is not None and not df_tabla.empty:
            for _, reg in df_tabla.iterrows():
                ws.cell(row=row, column=1, value=reg[col_etiqueta]).font = font_normal
                ws.cell(row=row, column=2, value=reg['gestionables']).font = font_normal
                ws.cell(row=row, column=3, value=reg['a_tiempo']).font = font_normal
                sla_cell = ws.cell(row=row, column=4, value=reg['sla'] / 100)
                sla_cell.number_format = '0.0%'
                sla_cell.font = font_normal
                for c in range(1, 5):
                    ws.cell(row=row, column=c).alignment = align_center
                if (row % 2) == 0:
                    for c in range(1, 5):
                        ws.cell(row=row, column=c).fill = fill_light
                row += 1
        else:
            ws.cell(row=row, column=1, value="Sin datos").font = font_normal
            row += 1

        if incluir_total and df_tabla is not None and not df_tabla.empty:
            total_gest = df_tabla['gestionables'].sum()
            total_tiempo = df_tabla['a_tiempo'].sum()
            sla_total = (total_tiempo / total_gest * 100) if total_gest > 0 else 0
            ws.cell(row=row, column=1, value="TOTAL GENERAL").font = font_bold
            ws.cell(row=row, column=2, value=total_gest).font = font_bold
            ws.cell(row=row, column=3, value=total_tiempo).font = font_bold
            sla_cell = ws.cell(row=row, column=4, value=sla_total / 100)
            sla_cell.number_format = '0.0%'
            sla_cell.font = font_bold
            for c in range(1, 5):
                ws.cell(row=row, column=c).alignment = align_center
            row += 1
        return row

    def agregar_grafico_combinado(ws, titulo_grafico, first_data_row, last_data_row,
                                  col_gest=2, col_sla=4, col_mes=1):
        if last_data_row < first_data_row:
            return
        chart = BarChart()
        chart.type = "col"
        chart.title = titulo_grafico
        chart.y_axis.title = "Gestionables"
        chart.style = 10

        data_gest = Reference(ws, min_col=col_gest, min_row=first_data_row, max_row=last_data_row)
        cats = Reference(ws, min_col=col_mes, min_row=first_data_row, max_row=last_data_row)
        chart.add_data(data_gest, titles_from_data=False)
        chart.set_categories(cats)
        chart.series[0].graphicalProperties.solidFill = "4f81bd"

        line = LineChart()
        data_sla = Reference(ws, min_col=col_sla, min_row=first_data_row, max_row=last_data_row)
        line.add_data(data_sla, titles_from_data=False)
        line.series[0].graphicalProperties.solidFill = "c00000"
        line.series[0].graphicalProperties.line.solidFill = "c00000"
        line.series[0].graphicalProperties.line.width = 25000
        line.series[0].marker.symbol = "circle"
        line.series[0].marker.size = 7
        line.y_axis.axId = 200
        line.y_axis.crosses = "max"
        chart.y_axis.crosses = "min"

        chart += line
        chart.legend.position = 'b'

        chart.series[0].dLbls = DataLabelList()
        chart.series[0].dLbls.showVal = True
        line.series[0].dLbls = DataLabelList()
        line.series[0].dLbls.numFmt = '0%'
        line.series[0].dLbls.showVal = True

        ws.add_chart(chart, f"A{last_data_row + 2}")

    # ----- Global mensual -----
    row_global = 2
    row_global = escribir_tabla(ws, row_global, "EVOLUCIÓN MENSUAL DEL SLA GLOBAL",
                                sla_global, C_HEADER_BLUE, incluir_total=True,
                                col_etiqueta='label', texto_encabezado='Mes')
    if sla_global is not None and not sla_global.empty:
        last_data_global = row_global - 2
        first_data_global = row_global - len(sla_global) - 1
        agregar_grafico_combinado(ws, "Evolución Mensual - Volumen y SLA Global",
                                  first_data_global, last_data_global)

    # ----- Zonas mensual -----
    if evo_zonas:
        zona_color = {'AMBA': C_TEAL, 'INTERIOR': C_ORANGE}
        for zona, df_zona in evo_zonas.items():
            current_row = row_global + 3
            color = zona_color.get(zona, C_BLUE)
            current_row = escribir_tabla(ws, current_row, f"SLA MENSUAL – ZONA {zona}",
                                         df_zona, color, incluir_total=True,
                                         col_etiqueta='label', texto_encabezado='Mes')
            if df_zona is not None and not df_zona.empty:
                first_data = current_row - len(df_zona) - 1
                last_data = current_row - 2
                agregar_grafico_combinado(ws, f"Evolución Mensual {zona} - Volumen y SLA",
                                          first_data, last_data)
            row_global = current_row

    # ----- Torta de distribución -----
    if evo_zonas and sla_global is not None and not sla_global.empty:
        totales = []
        etiquetas = []
        for zona, df_zona in evo_zonas.items():
            if df_zona is not None and not df_zona.empty:
                totales.append(df_zona['gestionables'].sum())
                etiquetas.append(zona)
        if totales:
            row_torta = row_global + 4 if row_global else 2
            ws.cell(row=row_torta, column=1, value="Categoría").font = font_header
            ws.cell(row=row_torta, column=2, value="Gestionables").font = font_header
            ws.cell(row=row_torta, column=1).fill = fill_header
            ws.cell(row=row_torta, column=2).fill = fill_header
            row_torta += 1
            for zona, total in zip(etiquetas, totales):
                ws.cell(row=row_torta, column=1, value=zona).font = font_normal
                ws.cell(row=row_torta, column=2, value=total).font = font_normal
                row_torta += 1
            pie = PieChart()
            data_pie = Reference(ws, min_col=2, min_row=row_torta - len(totales), max_row=row_torta - 1)
            cats_pie = Reference(ws, min_col=1, min_row=row_torta - len(totales), max_row=row_torta - 1)
            pie.add_data(data_pie, titles_from_data=False)
            pie.set_categories(cats_pie)
            pie.title = "Distribución de Envíos Gestionables"
            pie.series[0].dLbls = DataLabelList()
            pie.series[0].dLbls.showPercent = True
            pie.series[0].dLbls.showCatName = True
            ws.add_chart(pie, f"D{row_torta - len(totales)}")

    # ----- Hoja semanal aparte -----
    if sla_semanal is not None and not sla_semanal.empty:
        ws_sem = wb.create_sheet("SLA Semanal")
        ws_sem.column_dimensions['A'].width = 18
        ws_sem.column_dimensions['B'].width = 15
        ws_sem.column_dimensions['C'].width = 15
        ws_sem.column_dimensions['D'].width = 12

        row_sem = 2
        row_sem = escribir_tabla(ws_sem, row_sem, "CUMPLIMIENTO POR SEMANA CALENDARIO",
                                 sla_semanal, C_HEADER_BLUE, incluir_total=True,
                                 col_etiqueta='label', texto_encabezado='Semana')
        if not sla_semanal.empty:
            last_sem = row_sem - 2
            first_sem = row_sem - len(sla_semanal) - 1
            agregar_grafico_combinado(ws_sem, "Evolución Semanal - Volumen y SLA Global",
                                      first_sem, last_sem)

# ============================================================
# 4. FUNCIÓN PRINCIPAL DE PROCESAMIENTO (CACHEADA) - LÓGICA CORREGIDA
# ============================================================
@st.cache_data(show_spinner=False)
def procesar_datos(riders_bytes, riders_filename, zonas_bytes, zonas_filename):
    """
    Lee ambos archivos, cruza, excluye, calcula días hábiles e incumplimiento.
    Retorna (df, total_original, excluidas, total_encontradas).
    
    🔧 LÓGICA CORREGIDA: 
    - Órdenes ≤16:30: SLA de 3 días hábiles desde fecha de creación
    - Órdenes >16:30: SLA de 3 días hábiles desde el SIGUIENTE día hábil
    """
    df = pd.read_excel(io.BytesIO(riders_bytes), engine='openpyxl')
    df['NumeroGuia'] = df['NumeroGuia'].apply(pad_guia)
    df['FechaCreacionWMS'] = pd.to_datetime(df['FechaCreacionWMS'], errors='coerce')
    df['FechaEsperandoRetiro'] = pd.to_datetime(df['FechaEsperandoRetiro'], errors='coerce')

    # Eliminar Semana Calendario si existe (para evitar conflicto con el merge)
    if 'Semana Calendario' in df.columns:
        df = df.drop(columns=['Semana Calendario'])

    excluidas = 0
    total_encontradas = 0
    if zonas_bytes is not None:
        if zonas_filename.endswith('.csv'):
            df_zonas = pd.read_csv(io.BytesIO(zonas_bytes))
        else:
            df_zonas = pd.read_excel(io.BytesIO(zonas_bytes), engine='openpyxl')

        col_guia = None
        if 'NumeroGuia' in df_zonas.columns:
            col_guia = 'NumeroGuia'
        elif 'Guia' in df_zonas.columns:
            col_guia = 'Guia'

        if col_guia and 'ZONA' in df_zonas.columns:
            df_zonas[col_guia] = df_zonas[col_guia].apply(pad_guia)

            # Exclusión por destinatario (si existe la columna)
            tiene_dest = 'Destinatario' in df_zonas.columns
            if tiene_dest:
                mask = df_zonas['Destinatario'].astype(str).str.strip().str.contains(
                    'Supply Argentin', case=False, na=False
                )
                guias_excluir = set(df_zonas.loc[mask, col_guia].dropna().unique())
                total_encontradas = len(guias_excluir)
                if guias_excluir:
                    antes = len(df)
                    df = df[~df['NumeroGuia'].isin(guias_excluir)]
                    excluidas = antes - len(df)

            df_zonas = df_zonas.rename(columns={col_guia: 'NumeroGuia'})
            columnas_merge = ['NumeroGuia', 'ZONA']
            if tiene_dest:
                columnas_merge.append('Destinatario')
            if 'Semana Calendario' in df_zonas.columns:
                columnas_merge.append('Semana Calendario')

            df = df.merge(df_zonas[columnas_merge], on='NumeroGuia', how='left')

            if 'Semana Calendario' in df.columns:
                df['Semana Calendario'] = df['Semana Calendario'].apply(
                    lambda x: str(int(float(x))) if pd.notnull(x) and str(x).strip() != '' else ''
                )
        else:
            st.warning("⚠️ El archivo de zonas debe tener una columna de guía y 'ZONA'.")

    # Validar fechas incoherentes
    mask_inv = df['FechaEsperandoRetiro'].notna() & (df['FechaEsperandoRetiro'] < df['FechaCreacionWMS'])
    if mask_inv.any():
        st.warning(f"⚠️ {mask_inv.sum()} pedido(s) con FechaEsperandoRetiro anterior a creación fueron excluidos.")
        df = df[~mask_inv]

    total_original = len(df)

    # ============================================================
    # 🔧 CÁLCULO CORRECTO DE DÍAS HÁBILES SEGÚN CORTE 16:30
    # ============================================================
    
    # Días límite: SIEMPRE 3 días hábiles (el inicio del cómputo varía)
    df['dias_limite'] = 3
    
    # Hora de corte
    corte_hora = pd.Timestamp('1900-01-01 16:30:00').time()
    df['hora_creacion'] = df['FechaCreacionWMS'].dt.time
    
    # Cálculo vectorizado de feriados
    anios = df['FechaCreacionWMS'].dt.year.dropna().unique().astype(int).tolist()
    feriados_np = construir_feriados_np(anios)
    hoy = pd.Timestamp(datetime.now().replace(hour=0, minute=0, second=0, microsecond=0))
    
    # Fechas normalizadas para operaciones numpy
    fechas_creacion_np = df['FechaCreacionWMS'].dt.normalize().values.astype('datetime64[D]')
    fechas_fin = df['FechaEsperandoRetiro'].fillna(hoy).dt.normalize().values.astype('datetime64[D]')
    
    # 🔹 Determinar fecha de inicio del cómputo del SLA
    # Si orden ≤ 16:30: inicio = fecha de creación
    # Si orden > 16:30: inicio = siguiente día hábil (usando busday_offset)
    despues_del_corte = df['hora_creacion'] > corte_hora
    
    fecha_inicio_sla = np.where(
        despues_del_corte,
        np.busday_offset(fechas_creacion_np, 1, holidays=feriados_np, roll='forward'),  # siguiente hábil
        fechas_creacion_np  # misma fecha
    )
    
    # 🔹 Cálculo de días hábiles DESDE fecha_inicio_sla hasta FechaEsperandoRetiro
    df['dias_habiles'] = np.busday_count(fecha_inicio_sla, fechas_fin, holidays=feriados_np).clip(min=0)
    
    # 🔹 Incumplimiento: más de 3 días hábiles desde el inicio correcto del cómputo
    df['incumplimiento'] = df['dias_habiles'] > df['dias_limite']

    # ============================================================
    # 🔧 MOTIVO DETALLADO (CON LÓGICA CORREGIDA)
    # ============================================================
    def motivo_incumplimiento(row):
        dias = row['dias_habiles']
        limite = row['dias_limite']  # siempre 3
        creacion = row['FechaCreacionWMS']
        listo = row['FechaEsperandoRetiro']
        hora = row['hora_creacion']
        
        # Determinar si el SLA comenzó al día siguiente
        if hora <= corte_hora:
            inicio_sla = creacion.date()
            texto_inicio = f"desde creación ({creacion.date()})"
        else:
            # Buscar siguiente día hábil (para el mensaje)
            fecha_temp = creacion.date() + timedelta(days=1)
            while not es_dia_habil(fecha_temp):
                fecha_temp += timedelta(days=1)
            inicio_sla = fecha_temp
            texto_inicio = f"desde {inicio_sla} (orden post-corte 16:30 del {creacion.date()})"
        
        if not row['incumplimiento']:
            if pd.isna(listo):
                return f"Dentro del plazo pero aún no listo para retiro ({texto_inicio})"
            return f"Cumple ({texto_inicio})"
        
        if pd.isna(listo):
            return f"Aún no listo: {dias} días hábiles {texto_inicio} – supera los {limite} días permitidos"
        
        return f"Tomó {dias} días hábiles {texto_inicio}, listo {listo.date()} – supera los {limite} días permitidos"
    
    df['Motivo Incumplimiento'] = df.apply(motivo_incumplimiento, axis=1)

    return df, total_original, excluidas, total_encontradas

# ============================================================
# 5. APLICACIÓN PRINCIPAL STREAMLIT
# ============================================================
st.set_page_config(page_title="SLA Riders – Listo para Retiro", layout="wide")
st.title("🚚 SLA Riders – Pick Up Center (3 días hábiles desde inicio de preparación)")
st.markdown("""
**Regla de negocio:**  
- Pedidos creados **≤ 16:30** → preparación inicia mismo día → **3 días hábiles** desde creación.  
- Pedidos creados **> 16:30** → preparación inicia siguiente día hábil → **3 días hábiles** desde ese día.  
**La fecha de entrega NO se utiliza.** Si el pedido no tiene "FechaEsperandoRetiro", se considera aún no listo.
""")

archivo_riders = st.file_uploader("📂 Subí el archivo Excel", type=["xlsx"])
archivo_zonas = st.file_uploader("📂 (Opcional) Subí el archivo con ZONA, Destinatario y Semana", type=["xlsx", "csv"])

if archivo_riders is not None:
    # Convertir a bytes para cachear
    riders_bytes = archivo_riders.read()
    zonas_bytes = archivo_zonas.read() if archivo_zonas else None
    zonas_filename = archivo_zonas.name if archivo_zonas else ""

    # Procesamiento cacheado (se ejecuta una vez al cargar el archivo)
    df, total_original, excluidas, total_encontradas = procesar_datos(
        riders_bytes, archivo_riders.name, zonas_bytes, zonas_filename
    )

    # Mensaje de exclusión (sin duplicar el de ZONA)
    if archivo_zonas:
        if excluidas > 0:
            st.success(
                f"✅ Se añadió ZONA (y otras columnas) desde el archivo adicional.\n"
                f"🚫 Se excluyeron {excluidas} guías con destinatario 'Supply Argentin' "
                f"(se detectaron {total_encontradas} únicas)."
            )
        else:
            st.success("✅ Se añadió ZONA (y otras columnas) desde el archivo adicional.")

    # ============================================================
    # 🔧 SIDEBAR FILTROS CON FORMULARIO Y BOTÓN APLICAR
    # ============================================================
    st.sidebar.header("🔍 Filtros")

    # 📅 Fechas mínimas y máximas del dataset
    fecha_min = df['FechaCreacionWMS'].min().date()
    fecha_max = df['FechaCreacionWMS'].max().date()

    # 🎯 Filtro de zona (valores disponibles)
    opciones_zona = ['Todas']
    if 'ZONA' in df.columns:
        zonas_disponibles = df['ZONA'].dropna().unique().tolist()
        opciones_zona += sorted(zonas_disponibles)

    # 🔹 Formulario para agrupar filtros y evitar reprocesamiento en cada cambio
    with st.sidebar.form(key="form_filtros"):
        st.markdown("**📅 Filtro por fecha de creación**")
        col_fecha1, col_fecha2 = st.columns(2)
        
        with col_fecha1:
            fecha_desde_input = st.date_input(
                "Desde",
                value=fecha_min,
                min_value=fecha_min,
                max_value=fecha_max,
                format="MM/DD/YYYY"
            )
        
        with col_fecha2:
            fecha_hasta_input = st.date_input(
                "Hasta",
                value=fecha_max,
                min_value=fecha_min,
                max_value=fecha_max,
                format="MM/DD/YYYY"
            )

        zona_sel_input = st.selectbox("Zona", opciones_zona, key="zona_select")
        solo_incumplimientos_input = st.checkbox("Mostrar solo incumplimientos", value=False, key="incumpl_check")
        
        # 🔘 Botón para aplicar filtros (solo aquí se procesa el filtrado)
        submitted = st.form_submit_button("🔍 Aplicar filtros", type="primary")

    # 🔹 Si se presionó el botón, aplicar filtros; sino, usar valores por defecto
    if submitted:
        fecha_desde = fecha_desde_input
        fecha_hasta = fecha_hasta_input
        zona_sel = zona_sel_input
        solo_incumplimientos = solo_incumplimientos_input
    else:
        # Valores por defecto si no se aplicaron filtros aún
        fecha_desde = fecha_min
        fecha_hasta = fecha_max
        zona_sel = 'Todas'
        solo_incumplimientos = False

    # 🔹 Construir máscara de filtros
    mascara = pd.Series(True, index=df.index)
    mascara &= (df['FechaCreacionWMS'].dt.date >= fecha_desde) & \
               (df['FechaCreacionWMS'].dt.date <= fecha_hasta)
    
    if zona_sel != 'Todas' and 'ZONA' in df.columns:
        mascara &= (df['ZONA'] == zona_sel)
    if solo_incumplimientos:
        mascara &= df['incumplimiento']

    df_filtrado = df[mascara]   # copia solo las filas filtradas

    # Mostrar resumen de filtros aplicados
    st.sidebar.markdown("---")
    st.sidebar.markdown(f"**{len(df_filtrado):,}** pedidos con filtros aplicados")
    if submitted:
        st.sidebar.caption(f"📅 {fecha_desde} → {fecha_hasta} | 📍 {zona_sel} | {'🚫 Solo incumplimientos' if solo_incumplimientos else '✅ Todos'}")

    # ---------- MÉTRICAS ----------
    corte_hora = pd.Timestamp('1900-01-01 16:30:00').time()
    total_antes_1630 = (df_filtrado['hora_creacion'] <= corte_hora).sum()
    total_despues_1630 = len(df_filtrado) - total_antes_1630

    col1, col2, col3, col4 = st.columns(4)
    col1.metric("Total archivo", total_original)
    col2.metric("Creados ≤ 16:30 (SLA desde creación)", total_antes_1630)
    col3.metric("Creados > 16:30 (SLA desde día siguiente)", total_despues_1630)
    col4.metric("Órdenes únicas", df_filtrado['CodigoOrden'].nunique())

    # ---------- SLA GLOBAL ----------
    incumplidos = df_filtrado['incumplimiento'].sum()
    total_filtrado = len(df_filtrado)
    cumplidos = total_filtrado - incumplidos
    tasa_cumplimiento = (cumplidos / total_filtrado * 100) if total_filtrado else 0

    st.subheader("📊 Cumplimiento del SLA (objetivo ≥ 95%)")
    color_sla = "#28a745" if tasa_cumplimiento >= 95 else "#dc3545"
    color_inc = "#dc3545" if incumplidos > 0 else "#28a745"

    col_a, col_b = st.columns(2)
    with col_a:
        st.markdown(f"""
        <div style="background-color:{color_sla}18; border-left: 4px solid {color_sla};
                    padding: 12px 16px; border-radius: 6px;">
            <div style="font-size:13px; color:#555;">✅ Pedidos listos a tiempo</div>
            <div style="font-size:28px; font-weight:700; color:{color_sla};">{cumplidos:,}</div>
            <div style="font-size:14px; color:{color_sla};">{tasa_cumplimiento:.1f}%</div>
        </div>
        """, unsafe_allow_html=True)
    with col_b:
        st.markdown(f"""
        <div style="background-color:{color_inc}18; border-left: 4px solid {color_inc};
                    padding: 12px 16px; border-radius: 6px;">
            <div style="font-size:13px; color:#555;">🚫 Incumplimientos</div>
            <div style="font-size:28px; font-weight:700; color:{color_inc};">{incumplidos:,}</div>
            <div style="font-size:14px; color:{color_inc};">{(100 - tasa_cumplimiento):.1f}%</div>
        </div>
        """, unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)
    st.progress(min(tasa_cumplimiento / 100, 1.0))
    if tasa_cumplimiento >= 95:
        st.success(f"🎯 ¡Cumple el SLA del 95%! ({tasa_cumplimiento:.1f}%)")
    else:
        st.error(f"⚠️ No se alcanza el 95% – cumplimiento actual: {tasa_cumplimiento:.1f}%")

    # ============================================================
    # 🔧 FUNCIÓN AUXILIAR PARA EVOLUCIÓN - CORREGIDA DEFINITIVAMENTE
    # ============================================================
    def generar_evolucion(df, col_agrupar):
        """
        Genera tabla de evolución agrupada por mes o semana.
        CORRECCIÓN: Filtra valores vacíos y ordena semanas numéricamente.
        """
        if col_agrupar not in df.columns:
            return None
        
        # 🔹 Filtro robusto: elimina NaN, strings vacíos, espacios, y "nan" como texto
        def _es_valor_valido(x):
            if pd.isna(x):
                return False
            s = str(x).strip().lower()
            return s not in ['', 'nan', 'none', 'null']
        
        mask_valido = df[col_agrupar].apply(_es_valor_valido)
        df_valido = df[mask_valido].copy()
        
        if df_valido.empty:
            return None
            
        agg = df_valido.groupby(col_agrupar).agg(
            gestionables=('incumplimiento', 'count'),
            a_tiempo=('incumplimiento', lambda x: (~x).sum())
        ).reset_index()
        
        agg['sla'] = np.where(agg['gestionables'] > 0, 
                              agg['a_tiempo'] / agg['gestionables'] * 100, 0)
        
        # 🔹 ORDENAMIENTO CORRECTO
        if col_agrupar == 'Semana Calendario':
            # Convertir a numérico para orden cronológico real (15, 16, 17, 18...)
            agg['_orden_temp'] = pd.to_numeric(agg[col_agrupar], errors='coerce')
            # Eliminar filas donde la conversión falló (valores no numéricos residuales)
            agg = agg[agg['_orden_temp'].notna()].copy()
            if agg.empty:
                return None
            # Ordenar por valor numérico y limpiar columna temporal
            agg = agg.sort_values('_orden_temp').drop(columns=['_orden_temp']).reset_index(drop=True)
            # Generar label limpio para visualización
            agg['label'] = agg[col_agrupar].apply(lambda x: str(int(float(x))))
        else:  # 'mes' - orden cronológico natural de periodos YYYY-MM
            agg = agg.sort_values(col_agrupar).reset_index(drop=True)
            agg['label'] = pd.to_datetime(agg['mes'] + '-01').dt.strftime('%b %Y')
        
        return agg

    def aplicar_linea_95(fig):
        """Agrega línea de referencia al 95% en el eje Y secundario (SLA %)."""
        fig.add_hline(
            y=95,
            line_dash="dash",
            line_color="orange",
            line_width=2,
            annotation_text="Objetivo 95%",
            annotation_position="top right",
            annotation_font_color="orange",
            yref="y2"
        )
        return fig

    # ---------- EVOLUCIÓN MENSUAL ----------
    st.header("📅 Evolución del SLA")
    df_filtrado['mes'] = df_filtrado['FechaCreacionWMS'].dt.to_period('M').astype(str)

    sla_global_mes = generar_evolucion(df_filtrado, 'mes')
    sla_semanal = generar_evolucion(df_filtrado, 'Semana Calendario')

    evoluciones_por_zona = {}
    if 'ZONA' in df_filtrado.columns:
        for zona in ['AMBA', 'INTERIOR']:
            dfz = df_filtrado[df_filtrado['ZONA'] == zona]
            if not dfz.empty:
                evoluciones_por_zona[zona] = generar_evolucion(dfz, 'mes')

    # -- Visualización mensual global --
    if sla_global_mes is not None and not sla_global_mes.empty:
        st.subheader("🌍 SLA Global por Mes")
        col_tab, col_chart = st.columns([1, 2])
        with col_tab:
            st.dataframe(sla_global_mes[['label', 'gestionables', 'a_tiempo', 'sla']]
                         .rename(columns={'sla': 'SLA %'})
                         .assign(**{'SLA %': lambda x: x['SLA %'].map('{:.1f}%'.format)}), 
                         use_container_width=True)
        with col_chart:
            fig = px.bar(sla_global_mes, x='label', y='gestionables', text='gestionables',
                         title="Volumen y SLA Global", color_discrete_sequence=['#4f81bd'])
            fig.add_trace(go.Scatter(x=sla_global_mes['label'], y=sla_global_mes['sla'],
                                     mode='lines+markers', name='SLA %', yaxis='y2',
                                     line=dict(color='#c00000', width=3)))
            fig.update_layout(yaxis2=dict(overlaying='y', side='right', range=[0,100], ticksuffix='%'))
            fig = aplicar_linea_95(fig)
            st.plotly_chart(fig, use_container_width=True)

    # -- Visualización semanal -- 🔧 CORREGIDO: category_orders para ordenar eje X
    if sla_semanal is not None and not sla_semanal.empty:
        st.subheader("📅 Cumplimiento por Semana Calendario")
        col_tab2, col_chart2 = st.columns([1, 2])
        with col_tab2:
            st.dataframe(sla_semanal[['label', 'gestionables', 'a_tiempo', 'sla']]
                         .rename(columns={'sla': 'SLA %'})
                         .assign(**{'SLA %': lambda x: x['SLA %'].map('{:.1f}%'.format)}), 
                         use_container_width=True)
        with col_chart2:
            # 🔧 category_orders fuerza el orden correcto en el eje X para evitar saltos en la línea
            fig_sem = px.bar(sla_semanal, x='label', y='gestionables', text='gestionables',
                             title="Volumen y SLA Semanal", color_discrete_sequence=['#4f81bd'],
                             category_orders={'label': sla_semanal['label'].tolist()})
            fig_sem.add_trace(go.Scatter(x=sla_semanal['label'], y=sla_semanal['sla'],
                                         mode='lines+markers', name='SLA %', yaxis='y2',
                                         line=dict(color='#c00000', width=3)))
            fig_sem.update_layout(yaxis2=dict(overlaying='y', side='right', range=[0,100], ticksuffix='%'))
            fig_sem = aplicar_linea_95(fig_sem)
            st.plotly_chart(fig_sem, use_container_width=True)

    # -- Por zona mensual --
    if evoluciones_por_zona:
        st.subheader("📍 SLA por Zona (AMBA / INTERIOR)")
        tabs = st.tabs(list(evoluciones_por_zona.keys()))
        for zona, tab in zip(evoluciones_por_zona.keys(), tabs):
            with tab:
                sla_zona = evoluciones_por_zona[zona]
                if sla_zona is not None and not sla_zona.empty:
                    col1, col2 = st.columns([1, 2])
                    col1.dataframe(sla_zona[['label', 'gestionables', 'a_tiempo', 'sla']]
                                   .rename(columns={'sla': 'SLA %'})
                                   .assign(**{'SLA %': lambda x: x['SLA %'].map('{:.1f}%'.format)}), 
                                   use_container_width=True)
                    with col2:
                        fig_zona = px.bar(sla_zona, x='label', y='gestionables', text='gestionables',
                                         title=f"Evolución {zona}", 
                                         color_discrete_sequence=['#00b0b0' if zona=='AMBA' else '#ff9100'])
                        fig_zona.add_trace(go.Scatter(x=sla_zona['label'], y=sla_zona['sla'],
                                                      mode='lines+markers', name='SLA %', yaxis='y2',
                                                      line=dict(color='#c00000', width=3)))
                        fig_zona.update_layout(yaxis2=dict(overlaying='y', side='right', range=[0,100], ticksuffix='%'))
                        fig_zona = aplicar_linea_95(fig_zona)
                        st.plotly_chart(fig_zona, use_container_width=True)

    # ============================================================
    # 🚨 RANKING DE INCUMPLIMIENTOS - LÓGICA ORIENTADA AL NEGOCIO
    # ============================================================
    st.subheader("🚨 Períodos sin cumplir el objetivo (SLA < 95%)")

    col_rank1, col_rank2 = st.columns(2)

    def ranking_peores(df_base, col_agrupar, label_col, titulo):
        """
        Muestra períodos con SLA < 95% (incumplimientos reales).
        Si todos cumplen, muestra mensaje positivo.
        """
        if col_agrupar not in df_base.columns:
            return
        
        # 🔹 Filtro robusto de valores válidos
        def _es_valor_valido(x):
            if pd.isna(x):
                return False
            s = str(x).strip().lower()
            return s not in ['', 'nan', 'none', 'null']
        
        df_valido = df_base[df_base[col_agrupar].apply(_es_valor_valido)].copy()
        if df_valido.empty:
            return
            
        agg = df_valido.groupby(col_agrupar).agg(
            gestionables=('incumplimiento', 'count'),
            a_tiempo=('incumplimiento', lambda x: (~x).sum())
        ).reset_index()
        
        agg['sla'] = np.where(
            agg['gestionables'] > 0,
            agg['a_tiempo'] / agg['gestionables'] * 100, 0
        )
        
        # 🔹 Ordenamiento consistente por período
        if col_agrupar == 'Semana Calendario':
            agg['_orden_temp'] = pd.to_numeric(agg[col_agrupar], errors='coerce')
            agg = agg[agg['_orden_temp'].notna()]
            if agg.empty:
                return
            agg = agg.sort_values('_orden_temp').drop(columns=['_orden_temp'])
            agg['label'] = agg[col_agrupar].apply(lambda x: str(int(float(x))))
        else:
            agg = agg.sort_values(col_agrupar)
            agg['label'] = pd.to_datetime(agg[col_agrupar] + '-01').dt.strftime('%b %Y')
        
        # 🎯 LÓGICA ORIENTADA AL NEGOCIO:
        # 1. Filtrar solo los que NO cumplen el objetivo (<95%)
        incumplidores = agg[agg['sla'] < 95].copy()
        
        # 2. Preparar dataframe para mostrar
        resultado = incumplidores[['label', 'gestionables', 'a_tiempo', 'sla']].copy()
        
        # 3. Si hay incumplimientos: ordenar de PEOR a MEJOR (menor SLA primero)
        if not resultado.empty:
            resultado = resultado.sort_values('sla').reset_index(drop=True)
            resultado.columns = [label_col, 'Gestionables', 'A Tiempo', 'SLA %']
            resultado['SLA %'] = resultado['SLA %'].map('{:.1f}%'.format)
            
            # Limitar a 10 para no saturar (ajustable)
            if len(resultado) > 10:
                resultado = resultado.head(10)
                st.markdown(f"**{titulo}** (mostrando los 10 peores de {len(incumplidores)} incumplimientos)")
            else:
                st.markdown(f"**{titulo}** ({len(resultado)} período(s) sin cumplir el objetivo ≥95%)")
            
            st.dataframe(resultado, use_container_width=True)
        
        # 4. Si TODOS cumplen: mensaje positivo
        else:
            st.markdown(f"**{titulo}**")
            st.success(f"✅ Todos los períodos cumplen el objetivo de SLA ≥95%")

    with col_rank1:
        ranking_peores(df_filtrado, 'mes', 'Mes', '📅 Por Mes')
    with col_rank2:
        if 'Semana Calendario' in df_filtrado.columns:
            ranking_peores(df_filtrado, 'Semana Calendario', 'Semana', '📆 Por Semana')                    

    # ---------- TABLA DE DETALLE ----------
    st.subheader("📋 Detalle de pedidos analizados")

    columnas_mostrar = [
        'Operativa', 'Id', 'CodigoOrden', 'NumeroGuia',
        'FechaCreacionWMS', 'FechaEsperandoRetiro',
        'dias_limite', 'dias_habiles', 'incumplimiento', 'Motivo Incumplimiento'
    ]
    extras = []
    if 'ZONA' in df_filtrado.columns:
        extras.append('ZONA')
    if 'Destinatario' in df_filtrado.columns:
        extras.append('Destinatario')
    if 'Semana Calendario' in df_filtrado.columns:
        extras.append('Semana Calendario')
    columnas_mostrar = columnas_mostrar[:6] + extras + columnas_mostrar[6:]


    df_display = df_filtrado[columnas_mostrar].copy()
    if 'Id' in df_display.columns:
        df_display['Id'] = df_display['Id'].apply(
            lambda x: str(int(x)) if pd.notnull(x) and str(x).replace('.','',1).replace('-','').isdigit() else x
        )

    if 'Semana Calendario' in df_display.columns:
        df_display['Semana Calendario'] = df_display['Semana Calendario'].apply(
            lambda x: str(int(float(x))) if pd.notnull(x) and str(x).strip() != '' else ''
        )    

    def colorear_detalle(row):
        if row['incumplimiento'] == True:
            return ['background-color: #ffe5e5'] * len(row)
        return [''] * len(row)

    MAX_CELLS_STYLED = 262144
    total_celdas = len(df_display) * len(df_display.columns)

    if total_celdas <= MAX_CELLS_STYLED:
        st.dataframe(
            df_display.style.apply(colorear_detalle, axis=1),
            use_container_width=True
        )
    else:
        st.info(f"ℹ️ Se muestran {len(df_display):,} filas — coloreado desactivado por volumen. Descargá el Excel para ver el detalle completo.")
        st.dataframe(df_display, use_container_width=True)


    # ---------- DESCARGA EN EXCEL ----------
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Detalle SLA', index=False)
        escribir_hoja_evolucion(writer, sla_global_mes, evoluciones_por_zona, sla_semanal)
    output.seek(0)

    st.download_button(
        label="📥 Descargar resultados como Excel",
        data=output,
        file_name="sla_riders_completo.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )